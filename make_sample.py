"""
make_sample — 데모용 EPC 원가관리 엑셀을 코드로 생성한다.

포트폴리오/테스트를 위해 샘플 파일을 리포지터리에 바이너리로 넣지 않고,
누구나 재현할 수 있도록 스크립트로 만든다. 이 한 파일이 xlmeta의
모든 기능(업무 규칙 · 매칭 키 · 하드코딩 상수 · 수기 개입 · 소계 제외 ·
지표 의존 그래프)을 자극하도록 데이터를 구성했다.

    python make_sample.py            # → sample_epc_cost.xlsx
    python make_sample.py out.xlsx   # 경로 지정
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.workbook.defined_name import DefinedName


# ── 실적 시트 (원시 전표 데이터) ──────────────────────────────────
# 회사의 '진짜' 데이터. 결재상태·계정 같은 조건 컬럼이 여기 있고,
# 원가현황 시트의 SUMIFS가 이 표를 조건으로 훑는다.
LEDGER_HEADER = ["일자", "전표번호", "프로젝트코드", "계정", "거래처", "결재상태", "금액"]
LEDGER_ROWS = [
    ["2026-07-02", "V-0001", "P-2401", "직접비", "대성기계",   "승인", 320_000_000],
    ["2026-07-03", "V-0002", "P-2401", "간접비", "한빛엔지",   "승인",  45_000_000],
    ["2026-07-05", "V-0003", "P-2401", "직접비", "동원플랜트", "반려",  80_000_000],
    ["2026-07-06", "V-0004", "P-2402", "직접비", "대성기계",   "승인", 510_000_000],
    ["2026-07-08", "V-0005", "P-2402", "간접비", "삼우설비",   "대기",  30_000_000],
    ["2026-07-09", "V-0006", "P-2402", "직접비", "동원플랜트", "승인", 275_000_000],
    ["2026-07-10", "V-0007", "P-2403", "직접비", "한빛엔지",   "승인", 190_000_000],
    ["2026-07-11", "V-0008", "P-2403", "간접비", "삼우설비",   "승인",  62_000_000],
    ["2026-07-12", "V-0009", "P-2404", "직접비", "대성기계",   "승인", 430_000_000],
    ["2026-07-14", "V-0010", "P-2404", "직접비", "동원플랜트", "대기", 150_000_000],
    ["2026-07-15", "V-0011", "P-2405", "직접비", "삼우설비",   "승인", 210_000_000],
    ["2026-07-16", "V-0012", "P-2405", "간접비", "한빛엔지",   "반려",  25_000_000],
    ["2026-07-18", "V-0013", "P-2405", "직접비", "대성기계",   "승인",  95_000_000],
    ["2026-07-20", "V-0014", "P-2406", "직접비", "동원플랜트", "승인", 640_000_000],
    ["2026-07-22", "V-0015", "P-2406", "간접비", "삼우설비",   "승인",  88_000_000],
    ["2026-07-24", "V-0016", "P-2406", "직접비", "한빛엔지",   "대기",  70_000_000],
]

# ── 원가현황 시트 (담당자가 만든 요약 표) ────────────────────────
PROJECTS = [
    # 코드,     프로젝트명,       계약금액,        예산
    ["P-2401", "울산 정유 플랜트",  1_200_000_000,  1_000_000_000],
    ["P-2402", "서산 석유화학",     2_000_000_000,  1_800_000_000],
    ["P-2403", "여수 LNG 터미널",     800_000_000,    700_000_000],
    ["P-2404", "대산 정유 개보수",  1_500_000_000,  1_300_000_000],
    ["P-2405", "온산 발전 플랜트",    600_000_000,    550_000_000],
    ["P-2406", "광양 제철 설비",    1_100_000_000,  1_000_000_000],
]

# 수기 개입: P-2404(원가현황 9행) 발생원가는 수식이 아니라 사람이 박은 값.
MANUAL_ROW_INDEX = 3           # PROJECTS 상 P-2404
MANUAL_VALUE = 1_875_000_000
MANUAL_NOTE = "설계변경 미반영분 수기 가산 (6/28 PM 확인)"


def build_ledger(wb):
    ws = wb.create_sheet("실적")
    ws.append(LEDGER_HEADER)
    for c in range(1, len(LEDGER_HEADER) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for row in LEDGER_ROWS:
        ws.append(row)
    return ws


def build_cost_status(wb):
    ws = wb.create_sheet("원가현황")

    # 제목 행(4행): 셀 하나만 있어 layout이 '제목'으로 흡수한다.
    ws["A4"] = "원가현황 (2026년 7월)"
    ws["A4"].font = Font(bold=True, size=14)

    # 머리글 행(5행)
    header = ["프로젝트코드", "프로젝트명", "계약금액", "발생원가", "직접원가",
              "예산", "집행률", "예비비", "총투입예상", "비고"]
    for c, name in enumerate(header, start=1):
        cell = ws.cell(row=5, column=c, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 데이터 행(6~11행)
    for i, (code, name, contract, budget) in enumerate(PROJECTS):
        r = 6 + i
        ws.cell(row=r, column=1, value=code)                      # A 프로젝트코드
        ws.cell(row=r, column=2, value=name)                     # B 프로젝트명
        ws.cell(row=r, column=3, value=contract)                 # C 계약금액

        # D 발생원가: 승인된 전표만 합산 (업무 규칙이 여기 박제됨)
        if i == MANUAL_ROW_INDEX:
            ws.cell(row=r, column=4, value=MANUAL_VALUE)         # 수기 개입!
            ws.cell(row=r, column=10, value=MANUAL_NOTE)         # J 비고 = 사유
        else:
            ws.cell(row=r, column=4,
                    value=f'=SUMIFS(실적!G:G,실적!C:C,A{r},실적!F:F,"승인")')

        # E 직접원가: 승인 + 직접비 (업무 규칙 2개)
        ws.cell(row=r, column=5,
                value=f'=SUMIFS(실적!G:G,실적!C:C,A{r},실적!D:D,"직접비",실적!F:F,"승인")')
        ws.cell(row=r, column=6, value=budget)                   # F 예산
        ws.cell(row=r, column=7, value=f"=D{r}/F{r}")            # G 집행률
        ws.cell(row=r, column=8, value=f"=F{r}*0.05")            # H 예비비(상수 0.05)
        ws.cell(row=r, column=9, value=f"=D{r}+H{r}")            # I 총투입예상

    # 합계 행(12행): 소계/합계로 감지되어 지표 추출에서 제외되어야 한다.
    ws.cell(row=12, column=1, value="합계").font = Font(bold=True)
    for col in (3, 4, 5, 6, 8, 9):
        L = ws.cell(row=12, column=col).column_letter
        ws.cell(row=12, column=col, value=f"=SUM({L}6:{L}11)").font = Font(bold=True)

    return ws


# ── 경영요약 시트 (불일치 유발) ──────────────────────────────────
# 여기도 '발생원가'를 계산하지만, 결재상태(승인) 조건 없이 '전체'를 합한다.
# → 원가현황의 발생원가(승인만)와 정의가 달라 숫자가 안 맞는다.
#   xlmeta의 불일치 탐지가 "같은 '발생원가'가 두 곳에서 다르게 계산됨"을 잡는다.
def build_biz_summary(wb):
    ws = wb.create_sheet("경영요약")
    for c, name in enumerate(["프로젝트코드", "발생원가"], start=1):
        ws.cell(row=1, column=c, value=name).font = Font(bold=True)
    for i, code in enumerate(["P-2401", "P-2402", "P-2403", "P-2404", "P-2405"], start=2):
        ws.cell(row=i, column=1, value=code)
        # 발생원가: 결재상태 무관 '전체' 합 (원가현황은 승인만!) — 정의 불일치
        ws.cell(row=i, column=2, value=f'=SUMIFS(실적!G:G,실적!C:C,A{i})')
    return ws


def main(path="sample_epc_cost.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)               # 기본 시트 제거
    build_ledger(wb)
    build_cost_status(wb)
    build_biz_summary(wb)

    # 정의된 이름(named range): xlmeta가 name → 셀 로 뽑아 보고한다.
    wb.defined_names.add(DefinedName("실적원장", attr_text="실적!$A$1:$G$17"))
    wb.defined_names.add(DefinedName("프로젝트예산", attr_text="원가현황!$F$6:$F$11"))

    wb.save(path)
    print(f"생성: {path}  (시트: {[ws.title for ws in wb.worksheets]})")
    return path


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "sample_epc_cost.xlsx")
