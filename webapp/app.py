"""
xlmeta webapp — 엑셀을 올리면 xlmeta가 뽑아낸 지표 정의(OKF 번들)를
화면에서 탐색할 수 있게 하는 얇은 Flask 래퍼.

핵심 도구(xlmeta)는 그대로 두고, 여기서는 호출만 한다.

    pip install -r webapp/requirements.txt
    python webapp/app.py            # → http://127.0.0.1:5000
"""

import json
import os
import sys
import tempfile

from flask import Flask, Response, render_template, request

# 저장소 루트를 import 경로에 올려 xlmeta 패키지와 make_sample을 쓴다.
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HERE = os.path.dirname(os.path.abspath(__file__))
for p in (ROOT, HERE):
    if p not in sys.path:
        sys.path.insert(0, p)

from openpyxl import load_workbook                 # noqa: E402
from openpyxl.utils import get_column_letter        # noqa: E402

from xlmeta import extract, write_bundle          # noqa: E402
from xlmeta.emit_okf import slug                   # noqa: E402
from xlmeta.layout import analyze_sheet, cell_type  # noqa: E402
from xlmeta import formula as F                     # noqa: E402
from evaluate import Evaluator                       # noqa: E402
import make_sample                                 # noqa: E402

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024   # 10MB 상한
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0            # 정적파일 캐시 끔(개발용)
app.config["TEMPLATES_AUTO_RELOAD"] = True             # 템플릿 수정 즉시 반영


@app.after_request
def _no_cache(resp):
    resp.headers["Cache-Control"] = "no-store"
    return resp


# ── 구조 추출 (원본 격자 + 영역 역할) ──────────────────────────
MAX_ROWS, MAX_COLS = 120, 40


def _cell_display(v):
    """격자에 보여줄 문자열. 수식은 그대로, 숫자는 천단위로."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int) or (isinstance(v, float) and float(v).is_integer()):
        return f"{int(v):,}"
    if isinstance(v, float):
        return f"{v:,}"
    s = str(v)
    return s


def _role_of(reg, r, c):
    if r in reg.header_rows:
        return "header"
    if r in reg.subtotal_rows:
        return "subtotal"
    if c in reg.key_cols:
        return "key"
    return "data"


def build_structure(path):
    """시트별: 원본 격자(값·유형·구조역할·계산값) + 영역 요약 + 시트 참조."""
    wb = load_workbook(path, data_only=False)
    wbv = load_workbook(path, data_only=True)          # 엑셀이 저장한 캐시값(있으면)
    ev = Evaluator(wb)                                  # 없으면 직접 계산
    sheets = []
    for ws in wb.worksheets:
        regions = analyze_sheet(ws)
        H = min(ws.max_row or 1, MAX_ROWS)
        W = min(ws.max_column or 1, MAX_COLS)

        def region_at(r, c):
            for i, reg in enumerate(regions):
                if reg.r0 <= r <= reg.r1 and reg.c0 <= c <= reg.c1:
                    return i, reg
            return None, None

        grid = []
        for r in range(1, H + 1):
            row = []
            for c in range(1, W + 1):
                oc = ws.cell(row=r, column=c)
                v = oc.value
                is_f = oc.data_type == "f"           # 진짜 수식 여부
                gi, reg = region_at(r, c)
                cell = {
                    "v": _cell_display(v),
                    "t": cell_type(v, is_f),
                    "r": _role_of(reg, r, c) if reg else "out",
                    "g": gi,
                }
                if is_f:
                    cv = ev.value(ws.title, r, c)
                    if cv is None:                         # 캐시값으로 대체
                        cv = wbv[ws.title].cell(row=r, column=c).value
                    cell["cv"] = _cell_display(cv) if cv is not None else None
                row.append(cell)
            grid.append(row)

        reg_out = []
        for i, reg in enumerate(regions):
            tcell = (f"{get_column_letter(reg.title_cell[1])}{reg.title_cell[0]}"
                     if reg.title_cell else None)
            reg_out.append({
                "id": i,
                "a1": f"{ws.title}!{reg.a1}",
                "r0": reg.r0, "r1": reg.r1, "c0": reg.c0, "c1": reg.c1,
                "title": reg.title,
                "title_cell": tcell,        # 제목으로 추측한 셀 (A1 표기)
                "confidence": reg.confidence,
                "header_rows": reg.header_rows,
                "key_cols": [get_column_letter(c) for c in reg.key_cols],
                "subtotal_rows": reg.subtotal_rows,
                "columns": {get_column_letter(c): " > ".join(p)
                            for c, p in reg.col_paths.items()},
                "row_count": reg.r1 - reg.r0 + 1 - len(reg.header_rows),
            })
            # 격자에서 그 제목 셀에 표식을 남긴다 (영역 밖 위쪽 셀일 수 있음)
            if reg.title_cell:
                tr, tc = reg.title_cell
                if 1 <= tr <= H and 1 <= tc <= W:
                    grid[tr - 1][tc - 1]["title_of"] = i

        refs = set()
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    for s, _a in F.refs_in(cell.value[1:], ws.title):
                        if s != ws.title:
                            refs.add(s)

        sheets.append({
            "name": ws.title,
            "rows": H, "cols": W,
            "col_letters": [get_column_letter(c) for c in range(1, W + 1)],
            "truncated_rows": (ws.max_row or 0) > MAX_ROWS,
            "truncated_cols": (ws.max_column or 0) > MAX_COLS,
            "cells": grid,
            "regions": reg_out,
            "refs_out": sorted(refs),
        })
    return sheets


# ── 분석 파이프라인 ─────────────────────────────────────────────

def _md_key_for_metric(m):
    """emit_okf와 동일한 규칙으로 지표 → 마크다운 파일 경로를 만든다."""
    if m["confidence"]["level"] == "low":
        return None                                # 번들에서 제외됨
    title = m["title"] or m["id"]
    fn = slug(f"{title}-{m['id']}")
    return f"metrics/{fn}.md"


def _md_key_for_source(s):
    name = s["title"] or f"{s['sheet']} {s['range'].split('!')[1]}"
    return f"sources/{slug(name)}.md"


def analyze_path(xlsx_path):
    """엑셀 경로 → 화면이 쓸 JSON 묶음. 번들은 임시 폴더에 쓰고 읽어들인다."""
    meta = extract(xlsx_path)

    with tempfile.TemporaryDirectory() as outdir:
        stats = write_bundle(meta, outdir)
        bundle_md = {}
        for base, _dirs, files in os.walk(outdir):
            for fn in files:
                if fn.endswith(".md"):
                    full = os.path.join(base, fn)
                    rel = os.path.relpath(full, outdir).replace(os.sep, "/")
                    with open(full, encoding="utf-8") as f:
                        bundle_md[rel] = f.read()

    for m in meta["metrics"]:
        m["md_key"] = _md_key_for_metric(m)
    for s in meta["sources"]:
        s["md_key"] = _md_key_for_source(s)

    return {
        "source_file": meta["source_file"],
        "stats": {
            "metrics": stats["metrics"],
            "sources": stats["sources"],
            "excluded": stats["excluded"],
            "cells": stats["cells"],
            "unsupported": len(meta["unsupported"]),
        },
        "structure": build_structure(xlsx_path),
        "sources": meta["sources"],
        "metrics": meta["metrics"],
        "unsupported": meta["unsupported"],
        "bundle_md": bundle_md,
    }


def json_response(payload, status=200):
    """한글이 깨지지 않고 날짜 등도 안전하게 직렬화."""
    body = json.dumps(payload, ensure_ascii=False, default=str)
    return Response(body, status=status, mimetype="application/json")


# ── 라우트 ──────────────────────────────────────────────────────

@app.get("/")
def index():
    return render_template("index.html")


@app.post("/api/sample")
def api_sample():
    """내장 샘플 엑셀을 생성해 곧바로 분석한다 (파일 없이 체험용)."""
    try:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = os.path.join(tmp, "sample_epc_cost.xlsx")
            make_sample.main(xlsx)
            return json_response(analyze_path(xlsx))
    except Exception as e:                          # noqa: BLE001
        return json_response({"error": f"샘플 분석 실패: {e}"}, 500)


@app.post("/api/analyze")
def api_analyze():
    f = request.files.get("file")
    if f is None or not f.filename:
        return json_response({"error": "파일이 없습니다."}, 400)
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return json_response({"error": ".xlsx / .xlsm 파일만 지원합니다."}, 400)
    try:
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = os.path.join(tmp, os.path.basename(f.filename))
            f.save(xlsx)
            return json_response(analyze_path(xlsx))
    except Exception as e:                          # noqa: BLE001
        return json_response({"error": f"분석 실패: {e}"}, 500)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
