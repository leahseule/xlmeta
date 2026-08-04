"""
xlmeta.layout — 수식을 보기 전에 '표의 구조'를 먼저 확정한다.

순서:
  1) 셀을 값이 아니라 '유형'으로 바꾼 지도를 만든다
  2) 빈 행/열로 끊어 직사각형 영역(region)을 찾는다
  3) 영역마다 머리글 밴드를 찾는다 (텍스트만 있는 상단 행들)
  4) 병합셀을 채워 다단 머리글 경로를 만든다  예: 당월 > 실적
  5) 좌측 행머리글 열을 찾는다
  6) 소계/합계행을 걸러낸다
LLM을 쓰지 않는다. 못 정하면 None으로 남기고 confidence를 낮춘다.
"""

import re
from dataclasses import dataclass, field
from openpyxl.utils import get_column_letter

# 라벨 어디에 있어도 소계로 보는 단어 (충분히 길어 오인 위험이 낮음)
SUBTOTAL_WORDS = ("소계", "합계", "총계", "누계", "총합계", "subtotal", "total")
# 라벨 '전체'가 정확히 일치할 때만 소계로 보는 짧은 단어
# ("계"를 부분일치로 쓰면 '대성기계' 같은 거래처명이 소계로 오인됨)
SUBTOTAL_EXACT = ("계", "합", "sum")


# ── 1. 유형 지도 ────────────────────────────────────────────────

def cell_type(v, is_formula=False):
    # 수식 여부는 값이 '='로 시작하는지가 아니라 openpyxl data_type('f')로 판별한다.
    # ('=SUM(..)' 같은 아포스트로피 텍스트는 수식이 아니라 텍스트다.)
    if is_formula:
        return "f"
    if v is None or (isinstance(v, str) and not v.strip()):
        return "."          # 빈칸
    if isinstance(v, bool):
        return "b"
    if isinstance(v, (int, float)):
        return "n"          # 숫자
    return "t"              # 텍스트 ('='로 시작해도 수식이 아니면 텍스트)


def type_map(ws):
    h, w = ws.max_row, ws.max_column
    rows = []
    for r in range(1, h + 1):
        row = []
        for c in range(1, w + 1):
            cell = ws.cell(row=r, column=c)
            row.append(cell_type(cell.value, cell.data_type == "f"))
        rows.append(row)
    return rows, h, w


# ── 2. 영역 탐지 ────────────────────────────────────────────────

@dataclass
class Region:
    r0: int; r1: int; c0: int; c1: int          # 1-indexed 포함 범위
    header_rows: list = field(default_factory=list)
    key_cols: list = field(default_factory=list)
    col_paths: dict = field(default_factory=dict)   # col → ["당월","실적"]
    row_labels: dict = field(default_factory=dict)  # row → "울산 정유"
    subtotal_rows: list = field(default_factory=list)
    title: str = None
    confidence: str = "low"

    @property
    def a1(self):
        return f"{get_column_letter(self.c0)}{self.r0}:{get_column_letter(self.c1)}{self.r1}"


def find_regions(tmap, h, w, min_rows=2, min_cols=2):
    """빈 행/열을 경계로 연결된 덩어리를 찾는다 (flood fill)."""
    seen = [[False] * w for _ in range(h)]
    regions = []
    for i in range(h):
        for j in range(w):
            if seen[i][j] or tmap[i][j] == ".":
                continue
            stack, cells = [(i, j)], []
            seen[i][j] = True
            while stack:
                y, x = stack.pop()
                cells.append((y, x))
                for dy, dx in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    ny, nx = y + dy, x + dx
                    if 0 <= ny < h and 0 <= nx < w and not seen[ny][nx] and tmap[ny][nx] != ".":
                        seen[ny][nx] = True
                        stack.append((ny, nx))
            ys = [y for y, _ in cells]; xs = [x for _, x in cells]
            reg = Region(min(ys) + 1, max(ys) + 1, min(xs) + 1, max(xs) + 1)
            if (reg.r1 - reg.r0 + 1) >= min_rows and (reg.c1 - reg.c0 + 1) >= min_cols:
                regions.append(reg)
            else:
                reg.confidence = "fragment"
                regions.append(reg)
    return regions


# ── 3~4. 머리글 밴드 + 병합 해소 ─────────────────────────────────

def merged_lookup(ws):
    """병합 범위 안의 모든 좌표 → 앵커 값."""
    m = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                m[(r, c)] = v
    return m


def val(ws, mlk, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None and (r, c) in mlk:
        v = mlk[(r, c)]
    return v


def detect_header_band(ws, reg, tmap, mlk):
    """영역 상단에서 '텍스트 비중이 높고 아래가 숫자/수식으로 바뀌는' 행들."""
    band = []
    for r in range(reg.r0, min(reg.r1, reg.r0 + 4) + 1):
        types = [tmap[r - 1][c - 1] for c in range(reg.c0, reg.c1 + 1)]
        merged_here = sum(1 for c in range(reg.c0, reg.c1 + 1)
                          if (r, c) in mlk)
        n_text = types.count("t")
        n_data = types.count("n") + types.count("f")
        if n_data == 0 and (n_text >= 2 or merged_here >= 2):
            band.append(r)
        elif band:
            break
    return band


def build_col_paths(ws, reg, mlk):
    """다단 머리글을 경로로 조합. 상위가 병합이면 하위에 상속된다."""
    paths = {}
    for c in range(reg.c0, reg.c1 + 1):
        parts = []
        for r in reg.header_rows:
            v = val(ws, mlk, r, c)
            if isinstance(v, str) and v.strip():
                s = v.strip()
                if not parts or parts[-1] != s:
                    parts.append(s)
        if parts:
            paths[c] = parts
    return paths


# ── 5. 행머리글 열 ──────────────────────────────────────────────

def _is_id_column(ws, mlk, body, c):
    """숫자 열인데 값이 대부분 고유하면 ID/식별자로 본다.
       (금액·수량 같은 '집계형 데이터 숫자'는 값이 겹치므로 여기서 걸러진다.)"""
    nums = []
    for r in body:
        v = val(ws, mlk, r, c)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            nums.append(v)
    return len(nums) >= 2 and len(set(nums)) / len(nums) >= 0.9


def _short_text_col(ws, mlk, body, c, limit=16):
    """텍스트 열의 대푯값 길이가 짧으면 식별자(코드·명칭·범주),
       길면 내용(설명·비고)으로 본다. 중앙값으로 판단."""
    lens = [len(v.strip()) for r in body
            if isinstance(v := val(ws, mlk, r, c), str) and v.strip()]
    if not lens:
        return False
    lens.sort()
    return lens[len(lens) // 2] <= limit


def detect_key_cols(reg, tmap, ws, mlk):
    """행 식별 열 = 좌측의 식별자 열들.
       - 맨 앞 숫자 ID 열(고유값)도 인정한다 (예: 프로세스맵의 ID)
       - 텍스트 열은 짧은 식별자만 인정하고, 긴 설명 열이 나오면 멈춘다."""
    keys = []
    body = [r for r in range(reg.r0, reg.r1 + 1) if r not in reg.header_rows]
    for c in range(reg.c0, reg.c1 + 1):
        types = [tmap[r - 1][c - 1] for r in body]
        nonempty = [t for t in types if t != "."]
        if not nonempty:
            continue
        if types.count("t") / len(nonempty) >= 0.6:
            # 첫 식별 열은 무조건, 이후 텍스트 열은 '짧을 때만' 식별자로 인정
            if not keys or _short_text_col(ws, mlk, body, c):
                keys.append(c)
            else:
                break      # 긴 텍스트 = 내용 → 여기부터는 식별 열 아님
        elif not keys and _is_id_column(ws, mlk, body, c):
            keys.append(c)      # 맨 왼쪽 숫자 ID 열
        else:
            break          # 집계형 숫자 데이터나 그 오른쪽 → 멈춤
    return keys


# ── 6. 소계/합계행 ──────────────────────────────────────────────

def detect_subtotals(ws, reg, tmap, mlk):
    """라벨 패턴 또는 '같은 영역 위쪽을 SUM하는 수식'."""
    subs = []
    body = [r for r in range(reg.r0, reg.r1 + 1) if r not in reg.header_rows]
    for r in body:
        label = ""
        for c in reg.key_cols or [reg.c0]:
            v = val(ws, mlk, r, c)
            if isinstance(v, str):
                label += v.strip()
        lbl = label.strip()
        if any(w in lbl for w in SUBTOTAL_WORDS) or lbl in SUBTOTAL_EXACT:
            subs.append(r); continue
        for c in range(reg.c0, reg.c1 + 1):
            cell = ws.cell(row=r, column=c)
            if cell.data_type == "f":
                v = cell.value
                m = re.search(r"SUM\(\s*\$?[A-Z]{1,3}\$?(\d+)\s*:\s*\$?[A-Z]{1,3}\$?(\d+)\)",
                              v, re.I)
                if m and int(m.group(2)) < r and int(m.group(1)) >= reg.r0:
                    subs.append(r); break
    return sorted(set(subs))


# ── 제목 ────────────────────────────────────────────────────────

def absorb_title_row(ws, reg, tmap):
    """영역 첫 행에 셀이 하나뿐이고 아래 행은 여러 개면, 그 행은 제목이다."""
    while reg.r1 > reg.r0:
        first = [tmap[reg.r0 - 1][c - 1] for c in range(reg.c0, reg.c1 + 1)]
        below = [tmap[reg.r0][c - 1] for c in range(reg.c0, reg.c1 + 1)]
        if sum(1 for t in first if t != ".") == 1 and sum(1 for t in below if t != ".") >= 2:
            c = next(i for i, t in enumerate(first) if t != ".") + reg.c0
            v = ws.cell(row=reg.r0, column=c).value
            if isinstance(v, str):
                reg.title = v.strip()
            reg.r0 += 1
        else:
            break


def detect_title(ws, reg, tmap, regions):
    """영역 위쪽의 단독 텍스트 행. 다른 영역에 속한 행은 제외하고,
       후보가 여럿이면 글꼴이 가장 강조된(굵거나 큰) 것을 고른다."""
    if reg.title:
        return reg.title
    occupied = set()
    for o in regions:
        if o is not reg:
            occupied.update(range(o.r0, o.r1 + 1))
    best, best_score = None, -1
    for r in range(reg.r0 - 1, max(0, reg.r0 - 4), -1):
        if r < 1 or r in occupied:
            continue
        row_types = [tmap[r - 1][c - 1] for c in range(1, ws.max_column + 1)]
        if row_types.count("t") != 1 or any(t in "nf" for t in row_types):
            continue
        c = row_types.index("t") + 1
        if c > reg.c1 + 1:
            continue
        cell = ws.cell(row=r, column=c)
        if not isinstance(cell.value, str):
            continue
        f = cell.font
        score = (2 if f and f.bold else 0) + (f.size or 11) / 11.0
        if score > best_score:
            best, best_score = cell.value.strip(), score
    return best


# ── 조립 ────────────────────────────────────────────────────────

def analyze_sheet(ws):
    tmap, h, w = type_map(ws)
    mlk = merged_lookup(ws)
    regions = [r for r in find_regions(tmap, h, w) if r.confidence != "fragment"]

    for reg in regions:
        absorb_title_row(ws, reg, tmap)
        reg.header_rows = detect_header_band(ws, reg, tmap, mlk)
        reg.col_paths = build_col_paths(ws, reg, mlk)
        reg.key_cols = detect_key_cols(reg, tmap, ws, mlk)
        reg.subtotal_rows = detect_subtotals(ws, reg, tmap, mlk)
        reg.title = detect_title(ws, reg, tmap, regions)
        for r in range(reg.r0, reg.r1 + 1):
            if r in reg.header_rows or r in reg.subtotal_rows:
                continue
            lbl = " / ".join(str(val(ws, mlk, r, c)) for c in reg.key_cols
                             if val(ws, mlk, r, c) is not None)
            if lbl:
                reg.row_labels[r] = lbl
        data_rows = (reg.r1 - reg.r0 + 1) - len(reg.header_rows)
        covered = len(reg.col_paths) / max(1, reg.c1 - reg.c0 + 1)
        reg.confidence = ("high" if reg.header_rows and covered >= 0.9 and data_rows >= 2
                          else "medium" if reg.header_rows and covered >= 0.6
                          else "low")
    return regions


def describe(ws, regions):
    L = [f"# 시트: {ws.title}", ""]
    for i, reg in enumerate(regions, 1):
        L.append(f"## 영역 {i} — `{reg.a1}`  ({reg.confidence})")
        if reg.title:
            L.append(f"- 제목: {reg.title}")
        L.append(f"- 머리글 행: {reg.header_rows or '없음'}")
        L.append(f"- 행 식별 열: {[get_column_letter(c) for c in reg.key_cols] or '없음'}")
        if reg.subtotal_rows:
            L.append(f"- 소계/합계 행: {reg.subtotal_rows}  ← 지표 추출에서 제외")
        L.append("- 열 정의:")
        for c, path in reg.col_paths.items():
            L.append(f"    - `{get_column_letter(c)}` = {' > '.join(path)}")
        if reg.row_labels:
            head = list(reg.row_labels.items())[:3]
            L.append("- 행 라벨 예: " + ", ".join(f"{r}={v}" for r, v in head)
                     + (" …" if len(reg.row_labels) > 3 else ""))
        L.append("")
    return "\n".join(L)


if __name__ == "__main__":
    import sys
    from openpyxl import load_workbook
    wb = load_workbook(sys.argv[1])
    for ws in wb.worksheets:
        print(describe(ws, analyze_sheet(ws)))
