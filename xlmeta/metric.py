"""
xlmeta.metric — layout(구조)과 formula(수식)를 합쳐 지표를 만든다.
이 모듈만 둘 다 안다.
"""

import os
import re
from collections import defaultdict, OrderedDict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from . import formula as F
from . import explain as X
from . import insights as INS
from .evaluate import Evaluator
from .layout import analyze_sheet


def col_of(a1):
    m = re.match(r"^\$?([A-Z]{1,3})", a1)
    return column_index_from_string(m.group(1)) if m else None


def row_of(a1):
    m = re.match(r"^\$?[A-Z]{1,3}\$?(\d+)", a1)
    return int(m.group(1)) if m else None


class Book:
    """워크북 + 시트별 레이아웃 해석 결과를 함께 들고 다닌다."""

    def __init__(self, path):
        self.path = path
        self.wbf = load_workbook(path, data_only=False)
        self.wbv = load_workbook(path, data_only=True)
        self.layout = {ws.title: analyze_sheet(ws) for ws in self.wbf.worksheets}

    def region_at(self, sheet, row=None, col=None):
        """해당 좌표를 포함하는 영역. 행이 없으면 열만으로 찾는다."""
        for reg in self.layout.get(sheet, []):
            if col is not None and not (reg.c0 <= col <= reg.c1):
                continue
            if row is not None and not (reg.r0 <= row <= reg.r1):
                continue
            return reg
        return None

    def column_name(self, sheet, a1):
        """참조된 열의 이름을 레이아웃에서 조회. 예: 실적!G:G → '금액'."""
        c = col_of(a1)
        if c is None:
            return None, None
        r = row_of(a1)
        reg = self.region_at(sheet, r, c) or self.region_at(sheet, None, c)
        if not reg or c not in reg.col_paths:
            return None, None
        path = reg.col_paths[c]
        hdr_row = reg.header_rows[-1] if reg.header_rows else reg.r0
        return " > ".join(path), f"{sheet}!{get_column_letter(c)}{hdr_row}"


def build_metrics(book):
    """영역별로, 데이터 행에서만, 같은 수식 패턴이 반복되는 열을 지표로 삼는다."""
    metrics, cell_graph, unsupported = [], {}, []
    mid = 0

    for ws in book.wbf.worksheets:
        sheet = ws.title
        wsv = book.wbv[sheet]

        # 셀 단위 원장은 영역과 무관하게 전부 기록
        for row in ws.iter_rows():
            for c in row:
                if c.data_type == "f":               # 진짜 수식만 (아포스트로피 텍스트 제외)
                    body = c.value[1:]
                    cell_graph[f"{sheet}!{c.coordinate}"] = {
                        "formula": c.value,
                        "reads": [f"{s}!{a}" for s, a in F.refs_in(body, sheet)],
                        "value": wsv[c.coordinate].value,
                    }
                    for u in F.unsupported_in(body):
                        unsupported.append({"cell": f"{sheet}!{c.coordinate}",
                                            "reason": f"미지원 함수 {u}"})

        for reg in book.layout[sheet]:
            excluded = set(reg.header_rows) | set(reg.subtotal_rows)
            data_rows = [r for r in range(reg.r0, reg.r1 + 1) if r not in excluded]
            if not data_rows:
                continue

            blocks = defaultdict(list)
            for r in data_rows:
                for c in range(reg.c0, reg.c1 + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.data_type == "f":
                        blocks[(c, F.normalize(cell.value[1:], r))].append(r)

            for (c, pattern), rows in sorted(blocks.items(), key=lambda x: (x[0][0], min(x[1]))):
                rows.sort()
                anchor = rows[0]
                raw = ws.cell(row=anchor, column=c).value[1:]
                path = reg.col_paths.get(c)

                # 참조 해석
                reads = []
                for s, a in F.refs_in(raw, sheet):
                    name, hcell = book.column_name(s, a)
                    reads.append({"ref": f"{s}!{a}", "name": name, "name_cell": hcell})

                # 조건
                conds = F.extract_conditions(raw, sheet)
                for cd in conds:
                    s, _, a = cd["target_ref"].partition("!")
                    if a:
                        nm, _ = book.column_name(s, a)
                        cd["target_name"] = nm

                # 수기 개입: 데이터 행 중 이 열에 수식이 없는데 값이 있는 곳.
                # "수기값이다"는 결정론적으로 확실. '왜 넣었는지'는 추측하지 않는다.
                manual = []
                for r in data_rows:
                    if r < rows[0] or r > rows[-1]:
                        continue
                    oc = ws.cell(row=r, column=c)
                    v = oc.value
                    if v is not None and oc.data_type != "f":   # 값은 있는데 수식이 아님
                        manual.append({
                            "cell": f"{sheet}!{get_column_letter(c)}{r}",
                            "row_label": reg.row_labels.get(r),
                            "value": v,
                        })

                mid += 1
                metrics.append(OrderedDict([
                    ("id", f"M{mid:03d}"),
                    ("name", " > ".join(path) if path else None),
                    ("title", path[-1] if path else None),
                    ("sheet", sheet),
                    ("region", f"{sheet}!{reg.a1}"),
                    ("region_title", reg.title),
                    ("anchor_cell", f"{sheet}!{get_column_letter(c)}{anchor}"),
                    ("applies_to", f"{sheet}!{get_column_letter(c)}{rows[0]}"
                                   f":{get_column_letter(c)}{rows[-1]}"),
                    ("row_axis", [reg.row_labels.get(r) for r in rows]),
                    ("formula", "=" + raw),
                    ("pattern", pattern),
                    ("functions", F.functions_used(raw)),
                    ("reads", reads),
                    ("conditions", conds),
                    ("constants", F.hardcoded_constants(raw)),
                    ("manual_overrides", manual),
                    ("sample_value", wsv.cell(row=anchor, column=c).value),
                    ("confidence", _confidence(path, rows, manual, reg)),
                    ("_col", c),
                ]))

    _link(metrics)
    for m in metrics:
        m.pop("_col", None)
    return metrics, cell_graph, unsupported


def _confidence(path, rows, manual, reg):
    if not path:
        level = "low"                     # 이름을 못 정하면 신뢰할 수 없다
    elif len(rows) >= 3 and not manual and reg.confidence == "high":
        level = "high"
    elif len(rows) >= 2:
        level = "medium"
    else:
        level = "low"
    return {"level": level, "row_repeat": len(rows),
            "manual_exceptions": len(manual), "layout": reg.confidence}


def _link(metrics):
    """셀 참조를 지표 단위 의존으로 접어올린다."""
    index = {}
    for m in metrics:
        key = (m["sheet"], m["_col"])
        prev = index.get(key)
        if prev is None or m["confidence"]["row_repeat"] > prev[1]:
            index[key] = (m["id"], m["confidence"]["row_repeat"])
    by_id = {m["id"]: m for m in metrics}
    for m in metrics:
        dep = []
        for r in m["reads"]:
            s, _, a = r["ref"].partition("!")
            c = col_of(a)
            hit = index.get((s, c))
            if hit and hit[0] != m["id"] and hit[0] not in dep:
                dep.append(hit[0])
        m["depends_on"] = dep
        for d in dep:
            by_id[d].setdefault("used_by", [])
            if m["id"] not in by_id[d]["used_by"]:
                by_id[d]["used_by"].append(m["id"])
    for m in metrics:
        m.setdefault("used_by", [])


MAX_DATA_ROWS = 200


def _table_data(ev, sheet, reg):
    """표의 실제 데이터 행·값. 머리글 행은 빼고, 수식은 계산값으로 채운다
       (Evaluator가 수식은 계산, 일반 셀은 원값을 돌려줌)."""
    cols = list(range(reg.c0, reg.c1 + 1))
    col_letters = [get_column_letter(c) for c in cols]
    data_rows = [r for r in range(reg.r0, reg.r1 + 1) if r not in reg.header_rows]
    rows = [{"r": r, "cells": [ev.value(sheet, r, c) for c in cols]}
            for r in data_rows[:MAX_DATA_ROWS]]
    return {"col_letters": col_letters, "rows": rows,
            "total_rows": len(data_rows), "truncated": len(data_rows) > MAX_DATA_ROWS}


def extract(path):
    book = Book(path)
    metrics, cell_graph, unsupported = build_metrics(book)

    # 사람·AI가 읽을 지식 (설명·Python·함수 사전)
    for m in metrics:
        m["explanation"] = X.explain_metric(m)
        m["python"] = X.pythonize(m)
        m["functions_doc"] = X.functions_doc(m["functions"])

    ev = Evaluator(book.wbf)                 # 수식 계산값 산출 (데이터 값 채우기용)
    sources = []
    for sheet, regs in book.layout.items():
        for reg in regs:
            sources.append({
                "sheet": sheet,
                "range": f"{sheet}!{reg.a1}",
                "title": reg.title,
                "title_cell": (f"{get_column_letter(reg.title_cell[1])}{reg.title_cell[0]}"
                               if reg.title_cell else None),
                "columns": {get_column_letter(c): " > ".join(p)
                            for c, p in reg.col_paths.items()},
                "key_columns": [get_column_letter(c) for c in reg.key_cols],
                "header_rows": reg.header_rows,
                "subtotal_rows": reg.subtotal_rows,
                "row_count": reg.r1 - reg.r0 + 1 - len(reg.header_rows),
                "layout_confidence": reg.confidence,
                "data": _table_data(ev, sheet, reg),   # 실제 데이터 행·값 (수식은 계산값)
            })
    # 표(데이터 2행+)에 안 잡힌 비표 셀도 원문으로 — agent가 아무것도 놓치지 않게
    extra_cells = {}
    for ws in book.wbf.worksheets:
        real = [(reg.r0, reg.r1, reg.c0, reg.c1) for reg in book.layout[ws.title]
                if (reg.r1 - reg.r0 + 1 - len(reg.header_rows)) >= 2]
        cells = []
        for row in ws.iter_rows(max_row=min(ws.max_row or 1, 300), max_col=min(ws.max_column or 1, 60)):
            for cell in row:
                v = cell.value
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                if any(r0 <= cell.row <= r1 and c0 <= cell.column <= c1 for r0, r1, c0, c1 in real):
                    continue
                cells.append({"ref": cell.coordinate, "value": ev.value(ws.title, cell.row, cell.column)})
            if len(cells) >= 100:
                break
        if cells:
            extra_cells[ws.title] = cells[:100]

    p = book.wbf.properties
    _d = lambda x: x.strftime("%Y-%m-%d") if x else None
    properties = {
        "creator": p.creator or None,
        "last_modified_by": p.lastModifiedBy or None,
        "created": _d(p.created),
        "modified": _d(p.modified),
        "title": p.title or None,
    }

    named_ranges = []                        # 사용자 정의 이름만 (차트·인쇄영역 등 내부 이름 제외)
    _skip = ("Print_Area", "Print_Titles", "_FilterDatabase")
    try:
        for nm, dn in book.wbf.defined_names.items():
            if nm.startswith("_") or any(s in nm for s in _skip):
                continue                     # _xlchart·_xlfn·_xlnm·인쇄영역 = 노이즈
            val = getattr(dn, "value", None) or getattr(dn, "attr_text", None)
            named_ranges.append({"name": nm, "refers_to": str(val)})
    except Exception:                        # noqa: BLE001
        pass

    result = {
        "source_file": os.path.basename(path),
        "generated_by": "xlmeta 0.1 (no LLM)",
        "properties": properties,
        "named_ranges": named_ranges,
        "extra_cells": extra_cells,
        "sheets": [ws.title for ws in book.wbf.worksheets],
        "sources": sources,
        "metrics": metrics,
        "cell_graph": cell_graph,
        "unsupported": unsupported,
    }
    result["insights"] = INS.analyze(result)   # 의존 체인·순환·정의된 이름·불일치
    return result
