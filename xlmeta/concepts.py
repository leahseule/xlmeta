"""
xlmeta.concepts — 여러 시트에 흩어진 '같은 개념'을 하나의 의미 객체로 묶는다.

원칙: 추론하지 않는다. 작성자가 이미 남긴 링크(카테고리 코드 · 프로세스명 토큰)를
결정론적으로 파싱해 연결할 뿐이다. 글자가 전혀 안 겹치는 순수 의역은 여기서 잇지
않는다(그건 번들을 받은 LLM의 몫). 그래서 LLM이 필요 없다.

출력: Concept 객체 리스트. 각 객체는 belongs_to / 역할별 속성 / 연결된 아이디어·
진단 / evidence(시트!행)를 가진다.
"""

import re
from openpyxl.utils import get_column_letter

# ── 헤더 텍스트 → 의미 역할 (순서 = 우선순위, 먼저 맞는 게 이김) ─────────────
ROLE_KEYWORDS = [
    ("link",           ["연결 프로세스", "연결프로세스", "관련 프로세스", "연결", "관련"]),
    ("belongs_to",     ["밸류체인", "l1", "대분류", "카테고리", "영역", "분류"]),
    ("process",        ["l2 프로세스", "l2", "프로세스", "개념", "항목명"]),
    ("detail",         ["세부 업무", "세부업무", "세부"]),
    ("as_is",          ["as-is", "현행", "현재 방식"]),
    ("capture",        ["캡처", "capture"]),
    ("pain_point",     ["병목", "페인", "pain", "걸림돌", "리스크", "문제", "이슈"]),
    ("decision",       ["여기서 일어나는", "판단", "의사결정", "decision", "결정"]),
    ("ai_opportunity", ["ai 개입", "ai", "개입", "아이디어", "기회", "자동화"]),
    ("artifact",       ["산출", "문서·데이터", "산출 문서"]),
    ("owner",          ["담당", "owner", "책임"]),
    ("question",       ["진단 질문", "질문", "진단"]),
    ("finding",        ["알아내는", "알아내"]),
    ("signal",         ["신호", "이런 답"]),
]

# 개념 이름으로 부적절한 열(사람 이름 아님 등)은 seed에서 제외
_MIN_NAME_LEN = 2          # 정규화 후 최소 길이 (한글 기준)
_MAX_NAME_LEN = 24         # 이보다 길면 개념 이름이 아니라 문장으로 봄


def _norm(s):
    """공백·괄호·구분점 제거해 비교용 키로. '자동 발주'·'자동발주(B)' → '자동발주'."""
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"\([^)]*\)", "", s)                 # (B), (Assortment) 등 괄호 제거
    s = re.sub(r"[\s·/\-_,.\[\]()（）'\"★]+", "", s)  # 구분점·공백 제거
    return s.strip().lower()


def _role_of(header_text):
    h = str(header_text or "").strip().lower()
    for role, kws in ROLE_KEYWORDS:
        if any(k in h for k in kws):
            return role
    return None


def _cat_code(value):
    """'B. 수요예측·발주' → 'B'. 코드 없으면 None."""
    m = re.match(r"\s*([A-H])[.\s]", str(value or ""))
    return m.group(1) if m else None


def _parse_links(cell):
    """'발주(B), 마크다운(D)' → names={'발주','마크다운'}, codes={'B','D'}.
       'B 전체' / 'B ↔ G' → codes={'B'} / {'B','G'}."""
    text = str(cell or "")
    codes = set(re.findall(r"(?<![A-Za-z])([A-H])(?![A-Za-z가-힣])", text))
    names = set()
    for part in re.split(r"[,↔/]", text):        # 콤마 · ↔ · /
        m = re.match(r"\s*([^()]+?)\s*\(([A-H])\)", part)
        if m:
            names.add(m.group(1).strip())
    return names, codes


class _Table:
    """시트의 주 표: 1행=헤더, 이하=데이터. 레이아웃 과탐지에 안 기댄다."""

    def __init__(self, ws):
        self.title = ws.title
        self.header_row = 1
        self.cols = {}          # col_idx -> role
        self.headers = {}       # col_idx -> header text
        max_c = min(ws.max_column or 1, 40)
        for c in range(1, max_c + 1):
            h = ws.cell(self.header_row, c).value
            if h is None or not str(h).strip():
                continue
            self.headers[c] = str(h).strip()
            self.cols[c] = _role_of(h)
        self.ws = ws
        self.max_c = max_c

    def role_col(self, role):
        for c, r in self.cols.items():
            if r == role:
                return c
        return None

    def rows(self):
        ws = self.ws
        for r in range(self.header_row + 1, (ws.max_row or 1) + 1):
            vals = {c: ws.cell(r, c).value for c in self.headers}
            if sum(1 for v in vals.values() if v not in (None, "")) < 2:
                continue                              # 섹션 구분행 등은 건너뜀
            yield r, vals


def build_concepts(book):
    """Book(레이아웃 포함)에서 개념 객체 리스트를 만든다."""
    tables = [_Table(ws) for ws in book.wbv.worksheets]

    # 1) seed 표 = belongs_to + process 열을 모두 가진 표 (= 프로세스맵)
    seed_tabs = [t for t in tables
                 if t.role_col("belongs_to") is not None and t.role_col("process") is not None]
    if not seed_tabs:
        return []

    concepts = []       # 각: dict
    by_norm = {}        # norm(name) -> concept
    by_code = {}        # 카테고리 코드 -> [concept,...]

    for t in seed_tabs:
        c_name = t.role_col("process")
        c_cat = t.role_col("belongs_to")
        attr_cols = {role: c for c, role in t.cols.items()
                     if role in ("detail", "as_is", "pain_point", "decision",
                                 "ai_opportunity", "artifact", "owner")}
        for r, vals in t.rows():
            name = vals.get(c_name)
            if name is None or not str(name).strip():
                continue
            nn = _norm(name)
            if not (_MIN_NAME_LEN <= len(nn) <= _MAX_NAME_LEN):
                continue
            cat = vals.get(c_cat)
            code = _cat_code(cat)
            con = {
                "name": str(name).strip(),
                "norm": nn,
                "belongs_to": str(cat).strip() if cat else None,
                "code": code,
                "attrs": {},                 # role -> value
                "ideas": [],                 # 연결된 아이디어
                "diagnostics": [],           # 연결된 진단 질문
                "evidence": [f"{t.title}!row{r}"],
            }
            for role, c in attr_cols.items():
                v = vals.get(c)
                if v not in (None, ""):
                    con["attrs"][role] = str(v).strip()
            concepts.append(con)
            by_norm.setdefault(nn, con)
            if code:
                by_code.setdefault(code, []).append(con)

    # 2) 링크 표(나머지)에서 아이디어·진단을 개념에 붙인다
    def _match_by_name(token):
        """토큰 → (정확일치 개념들, 부분일치 개념들). 정확=high, 부분=medium."""
        tn = _norm(token)
        if len(tn) < _MIN_NAME_LEN:
            return [], []
        if tn in by_norm:
            return [by_norm[tn]], []
        subs = [c for n, c in by_norm.items()
                if len(tn) >= 3 and (tn in n or n in tn)]
        return [], subs

    for t in tables:
        if t in seed_tabs:
            continue
        c_link = t.role_col("link")
        c_area = t.role_col("belongs_to")       # 진단질문 '영역'
        c_idea = t.role_col("ai_opportunity")
        c_risk = t.role_col("pain_point")
        c_q = t.role_col("question")
        c_sig = t.role_col("signal")
        c_find = t.role_col("finding")
        # 텍스트 스캔 대상 열(개념명이 문장 안에 박혀 있을 수 있음)
        text_cols = [c for c, role in t.cols.items()
                     if role in ("ai_opportunity", "question", "finding", "decision")]

        for r, vals in t.rows():
            targets = {}          # id(concept) -> (concept, confidence)
            _rank = {"high": 2, "medium": 1}

            def _add(cons, conf):
                for c in cons:
                    prev = targets.get(id(c))
                    if prev is None or _rank[conf] > _rank[prev[1]]:
                        targets[id(c)] = (c, conf)

            # (a) 명시적 링크 토큰 (연결 프로세스) — 이름만 사용. 코드-only는
            #     영역 전체를 뭉뚱그려 모든 개념에 노이즈를 뿌리므로 붙이지 않는다.
            if c_link is not None:
                names, _codes = _parse_links(vals.get(c_link))
                for nm in names:
                    ex, sub = _match_by_name(nm)
                    _add(ex, "high")
                    _add(sub, "medium")
            # (b) 영역 열이 코드가 아니라 이름이면 개념 매칭 (코드면 무시)
            if c_area is not None:
                area = vals.get(c_area)
                if area and not _cat_code(area):
                    ex, sub = _match_by_name(area)
                    _add(ex, "high")
                    _add(sub, "medium")
            # (c) 본문 텍스트에 개념명이 박혀 있으면 (부분일치, 3자 이상)
            for c in text_cols:
                txt = _norm(vals.get(c))
                if not txt:
                    continue
                for n, con in by_norm.items():
                    if len(n) >= 3 and n in txt:
                        _add([con], "medium")

            if not targets:
                continue
            ev = f"{t.title}!row{r}"
            idea_txt = vals.get(c_idea) if c_idea else None
            risk_txt = vals.get(c_risk) if c_risk else None
            q_txt = vals.get(c_q) if c_q else None
            sig_txt = vals.get(c_sig) if c_sig else None
            find_txt = vals.get(c_find) if c_find else None

            for con, conf in targets.values():
                if idea_txt:
                    con["ideas"].append({
                        "text": str(idea_txt).strip(),
                        "risk": str(risk_txt).strip() if risk_txt else None,
                        "where": ev, "confidence": conf,
                    })
                if q_txt:
                    con["diagnostics"].append({
                        "question": str(q_txt).strip(),
                        "finding": str(find_txt).strip() if find_txt else None,
                        "signal": str(sig_txt).strip() if sig_txt else None,
                        "where": ev, "confidence": conf,
                    })
                if ev not in con["evidence"]:
                    con["evidence"].append(ev)

    # 3) 연결이 하나도 없는(단독) 개념은 의미 객체로서 가치가 적으니 뒤로
    for con in concepts:
        con["links"] = len(con["ideas"]) + len(con["diagnostics"])
        con.pop("norm", None)
    concepts.sort(key=lambda c: -c["links"])
    return concepts


# ── 마크다운 렌더 (OKF · 요약 · 프리필 공용) ──────────────────────────────
_ATTR_LABEL = {
    "decision": "판단", "pain_point": "페인포인트", "as_is": "현행",
    "capture": "캡처 가능성", "owner": "담당", "detail": "세부업무",
    "artifact": "산출물",
}
_CONF = {"high": "확실", "medium": "추정"}


def concepts_md(concepts):
    """개념 객체 → 마크다운. 교차연결된 개념은 상세히, 단독은 목록으로."""
    if not concepts:
        return ""
    linked = [c for c in concepts if c["links"] > 0]
    solo = [c for c in concepts if c["links"] == 0]
    L = ["## 개념(Concept) — 여러 시트에 흩어진 같은 개념을 하나로", ""]
    L.append("> 같은 개념이 프로세스·아이디어·진단 시트에 나뉘어 있던 것을 한 객체로 "
             "조립했습니다. 근거 위치(`시트!행`)를 함께 답니다. LLM 없이 작성자가 남긴 "
             "연결 토큰·이름 일치로만 이었습니다.")
    L.append("")
    for con in linked:
        L.append(f"### {con['name']}")
        if con.get("belongs_to"):
            L.append(f"- **소속**: {con['belongs_to']}")
        for role in ("decision", "pain_point", "as_is", "owner"):
            v = con["attrs"].get(role)
            if v:
                L.append(f"- **{_ATTR_LABEL[role]}**: {v}")
        if con["ideas"]:
            L.append("- **AI 기회**:")
            for it in con["ideas"]:
                risk = f" · 리스크: {it['risk']}" if it.get("risk") else ""
                L.append(f"  - [{_CONF.get(it['confidence'],'')}] {it['text']}{risk} `{it['where']}`")
        if con["diagnostics"]:
            L.append("- **진단 질문**:")
            for d in con["diagnostics"]:
                sig = f" → {d['signal']}" if d.get("signal") else ""
                L.append(f"  - [{_CONF.get(d['confidence'],'')}] {d['question']}{sig} `{d['where']}`")
        L.append(f"- 근거: {', '.join('`'+e+'`' for e in con['evidence'])}")
        L.append("")
    if solo:
        names = ", ".join(f"{c['name']}"
                          + (f"({c['code']})" if c.get("code") else "") for c in solo)
        L.append(f"**단독 개념** (교차 연결 없음, {len(solo)}개): {names}")
        L.append("")
    return "\n".join(L)
